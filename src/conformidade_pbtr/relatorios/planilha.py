"""Relatório em XLSX — checklist tabular para acompanhamento das pendências."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..modelos import Categoria, Relatorio, Status

PREENCHIMENTO = {
    Status.CONFORME: "C6EFCE",
    Status.NAO_CONFORME: "FFC7CE",
    Status.ATENCAO: "FFEB9C",
    Status.VERIFICAR_MANUAL: "DDEBF7",
    Status.NAO_APLICAVEL: "F2F2F2",
}

ROTULO = {
    Status.CONFORME: "Conforme",
    Status.NAO_CONFORME: "Não conforme",
    Status.ATENCAO: "Atenção",
    Status.VERIFICAR_MANUAL: "Verificar manualmente",
    Status.NAO_APLICAVEL: "Não aplicável",
}

CABECALHO = Font(bold=True, color="FFFFFF", size=10)
FILL_CAB = PatternFill("solid", fgColor="1F4E79")
BORDA = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _ajustar(ws, larguras: list[int]) -> None:
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def gerar_xlsx(rel: Relatorio, destino: str | Path) -> str:
    d = Path(destino)
    wb = Workbook()

    # ---------------------------------------------------------- resumo
    ws = wb.active
    ws.title = "Resumo"
    r = rel.resumo
    linhas = [
        ("Documento", rel.documento.nome),
        ("Tipo", rel.documento.tipo),
        ("Formato analisado", f".{rel.documento.formato}"),
        ("Páginas", rel.documento.paginas),
        ("Tabelas detectadas", len(rel.documento.tabelas)),
        ("Contexto inferido", ", ".join(rel.documento.tags_contexto)),
        ("Checklist", f"v{rel.versao_checklist}"),
        ("Gerado em", rel.gerado_em),
        ("", ""),
        ("Índice de conformidade (%)", r.indice_conformidade),
        ("Conforme", r.conforme),
        ("Não conforme", r.nao_conforme),
        ("Atenção", r.atencao),
        ("Verificar manualmente", r.verificar_manual),
        ("Não aplicável", r.nao_aplicavel),
        ("", ""),
        ("Erros de numeração", r.erros_numeracao),
        ("Divergências em tabelas/valores", r.erros_tabela),
        ("Apontamentos de revisão textual (regra)", r.erros_ortografia),
        ("Sugestões da revisão pelo agente", r.sugestoes_ia),
    ]
    ws["A1"] = "RELATÓRIO DE CONFORMIDADE — PB/TR"
    ws["A1"].font = Font(bold=True, size=13)
    for i, (k, v) in enumerate(linhas, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    _ajustar(ws, [34, 60])

    # ------------------------------------------------------- checklist
    ws2 = wb.create_sheet("Checklist")
    cab = [
        "ID", "Seção", "Item do roteiro", "Severidade", "Situação",
        "Constatação", "Recomendação", "Item", "Pág.", "Responsável", "Prazo", "Status do tratamento",
    ]
    ws2.append(cab)
    for c in ws2[1]:
        c.font = CABECALHO
        c.fill = FILL_CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for a in sorted(rel.por_categoria(Categoria.CHECKLIST), key=lambda x: x.id):
        ws2.append([
            a.id, a.secao, a.titulo, a.severidade.value, ROTULO[a.status],
            a.encontrado, a.orientacao, a.item, a.pagina or "", "", "", "Pendente",
        ])
        ws2.cell(row=ws2.max_row, column=5).fill = PatternFill(
            "solid", fgColor=PREENCHIMENTO[a.status]
        )

    dv = DataValidation(
        type="list",
        formula1='"Pendente,Em tratamento,Resolvido,Não se aplica"',
        allow_blank=True,
    )
    ws2.add_data_validation(dv)
    dv.add(f"L2:L{max(ws2.max_row, 2)}")

    for linha in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=len(cab)):
        for c in linha:
            c.border = BORDA
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cab))}{ws2.max_row}"
    _ajustar(ws2, [12, 26, 42, 12, 16, 40, 46, 9, 7, 16, 12, 18])

    # ----------------------------------------- verificações automáticas
    ws3 = wb.create_sheet("Automáticas")
    cab3 = ["ID", "Categoria", "Origem", "Título", "Situação", "Severidade",
            "Esperado", "Encontrado", "Trecho", "Item", "Pág.", "Recomendação"]
    ws3.append(cab3)
    for c in ws3[1]:
        c.font = CABECALHO
        c.fill = FILL_CAB
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for cat in (Categoria.ESTRUTURA, Categoria.NUMERACAO, Categoria.TABELA,
                Categoria.VALOR, Categoria.ORTOGRAFIA):
        for a in rel.por_categoria(cat):
            ws3.append([
                a.id, cat.value,
                "IA (sugestão)" if a.origem.value == "ia" else "determinística",
                a.titulo, ROTULO[a.status], a.severidade.value,
                a.esperado, a.encontrado, a.evidencia[:250], a.item, a.pagina or "", a.orientacao,
            ])
            ws3.cell(row=ws3.max_row, column=5).fill = PatternFill(
                "solid", fgColor=PREENCHIMENTO[a.status]
            )
    for linha in ws3.iter_rows(min_row=1, max_row=ws3.max_row, max_col=len(cab3)):
        for c in linha:
            c.border = BORDA
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(cab3))}{ws3.max_row}"
    _ajustar(ws3, [14, 13, 15, 40, 16, 12, 28, 28, 36, 9, 7, 38])

    wb.save(str(d))
    return str(d)
